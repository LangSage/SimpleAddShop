from django.shortcuts import render, get_object_or_404
from .models import Category, SubCategory, Product, Slider
from .forms import ProductSearchForm
from django.db.models import Q
from blog.models import Post  # Assuming the blog app has a Post model
from django.contrib.auth.decorators import user_passes_test


def index(request):
    slider     = Slider.objects.filter(is_active=True).first()
    categories = Category.objects.all()
    posts      = Post.objects.order_by('-published_at')[:3]
    return render(request, 'shop/index.html', {
        'slider':     slider,
        'categories': categories,
        'posts':      posts,
    })

def category_list(request):
    categories = Category.objects.all()
    return render(request, 'shop/category_list.html', {'categories': categories})


def product_search(request):
    form = ProductSearchForm(request.GET)
    query = request.GET.get('query')
    results = []

    if query:
        results = Product.objects.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query)
        )

    # Fetch all categories to display in the sidebar
    categories = Category.objects.all()

    context = {
        'form': form,
        'results': results,
        'query': query,
        'categories': categories,  # Include categories in the context
    }

    return render(request, 'shop/product_search.html', context)



def category_detail(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    subcategories = category.subcategories.all()
    products = Product.objects.filter(
        category=category
    ).exclude(
        subcategory__in=subcategories
    )
    return render(request, 'shop/category_detail.html', {
        'category': category,
        'subcategories': subcategories,
        'products': products,
        'categories': Category.objects.all()
    })

def subcategory_list(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    subcategories = category.subcategories.all()
    products_in_subcategories = Product.objects.filter(subcategory__in=subcategories)
    products = Product.objects.filter(
        category=category
    ).exclude(
        subcategory__in=subcategories
    ) | products_in_subcategories
    return render(request, 'shop/subcategory_list.html', {
        'category': category,
        'subcategories': subcategories,
        'products': products,
        'categories': Category.objects.all()
    })

def product_list(request, subcategory_id):
    subcategory = get_object_or_404(SubCategory, id=subcategory_id)
    products = Product.objects.filter(subcategory=subcategory)
    return render(request, 'shop/product_list.html', {
        'subcategory': subcategory,
        'products': products,
        'categories': Category.objects.all()
    })

def product_detail(request, slug):
    product = get_object_or_404(Product, slug=slug)
    return render(request, 'shop/product_detail.html', {
        'product': product,
        'categories': Category.objects.all()
    })



def about_company(request):
    # Fetch all categories
    categories = Category.objects.all()
    # Pass categories to the template
    return render(request, 'shop/about_company.html', {'categories': categories})


from django.shortcuts import render, redirect, get_object_or_404
from .forms import OrderForm
from .models import Product, Order
from django.contrib import messages

def order_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            Order.objects.create(
                product=product,
                name=form.cleaned_data['name'],
                email=form.cleaned_data['email'],
                phone=form.cleaned_data['phone'],
                message=form.cleaned_data['message'],
            )
            messages.success(request, "Ваш заказ успешно оформлен!")
            return redirect(request.META.get('HTTP_REFERER', '/'))
    else:
        form = OrderForm()

    return render(request, 'shop/order_form.html', {'form': form, 'product': product})


import csv
from django.http import HttpResponse
from .models import Product
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Product
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def import_export_page(request):
    return render(request, 'shop/import_export.html')

@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def export_products_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ['ID', 'Name', 'Category', 'Subcategory', 'Meta Title', 'Meta Description', 'Meta Keywords']
    ws.append(headers)

    for product in Product.objects.all():
        ws.append([
            product.id,
            product.name,
            product.category.name if product.category else '',
            product.subcategory.name if product.subcategory else '',
            product.meta_title or '',
            product.meta_description or '',
            product.meta_keywords or ''
        ])

    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
    return response


from django.contrib import messages
import openpyxl
@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def import_products_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "❌ Файл не был загружен.")
            return redirect('import_export_page')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"❌ Не удалось прочитать файл Excel: {str(e)}")
            return redirect('import_export_page')

        headers = [cell.value for cell in sheet[1]]
        updated_log = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            product_id = row_data.get('ID')
            if not product_id:
                continue
            try:
                product = Product.objects.get(id=product_id)
                changes = []

                # Проверяем каждое поле и логируем изменение
                if row_data.get('Name') and row_data['Name'] != product.name:
                    changes.append(f"Name: \"{product.name}\" → \"{row_data['Name']}\"")
                    product.name = row_data['Name']

                if row_data.get('Meta Title') != product.meta_title:
                    changes.append(f"Meta Title: \"{product.meta_title}\" → \"{row_data.get('Meta Title', '')}\"")
                    product.meta_title = row_data.get('Meta Title', '')

                if row_data.get('Meta Description') != product.meta_description:
                    changes.append(f"Meta Description: \"{product.meta_description}\" → \"{row_data.get('Meta Description', '')}\"")
                    product.meta_description = row_data.get('Meta Description', '')

                if row_data.get('Meta Keywords') != product.meta_keywords:
                    changes.append(f"Meta Keywords: \"{product.meta_keywords}\" → \"{row_data.get('Meta Keywords', '')}\"")
                    product.meta_keywords = row_data.get('Meta Keywords', '')

                if changes:
                    product.save()
                    updated_log.append(f"🛠 <b>{product.name}</b> (ID {product.id}):<br> " + "<br> ".join(changes))

            except Product.DoesNotExist:
                messages.warning(request, f"⚠️ Товар с ID {product_id} не найден.")
                continue

        if updated_log:
            messages.success(request, f"✅ Изменения внесены в {len(updated_log)} товаров.")
            for log in updated_log[:10]:  # показываем максимум 10
                messages.info(request, log)
        else:
            messages.info(request, "🔍 Все поля уже были актуальны, ничего не изменено.")

        return redirect('import_export_page')




# ________________Slider on main page____________________________


# ________________Slider on main page____________________________